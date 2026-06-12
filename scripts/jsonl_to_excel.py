#!/usr/bin/env python3
"""
JSONL → Excel（三 Sheet 格式）

  Sheet 1: 认缴流量 (subscription_flow)  — 每行=一个认购
  Sheet 2: 股权存量 (equity_snapshot)   — 每行=一个股东持仓
  Sheet 3: schema_cross_check           — Pydantic 校验 + Cross-check 结果

输出: outputs/week2_excel/{code}_{name}_三表抽取.xlsx
"""
import sys
import json
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_XL = True
except ImportError:
    HAS_XL = False

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

JSONL_DIR = PROJECT_ROOT / "outputs" / "week2_jsonl"
EXCEL_DIR = PROJECT_ROOT / "outputs" / "week2_excel"
EXCEL_DIR.mkdir(parents=True, exist_ok=True)

# 样式
HF = Font(bold=True, size=11)
HFILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
HA = Alignment(horizontal="center", vertical="center", wrap_text=True)
CA = Alignment(vertical="top", wrap_text=True)
BDR = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))


def _header(ws, headers):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HF; cell.fill = HFILL; cell.alignment = HA; cell.border = BDR


def _widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def jsonl_to_excel(path, out_path):
    with open(path, encoding="utf-8") as f:
        rows = [json.loads(ln) for ln in f if ln.strip()]

    sub_flows = [r for r in rows if r.get("record_type") == "subscription_flow"]
    eq_snaps = [r for r in rows if r.get("record_type") == "equity_snapshot"]

    wb = openpyxl.Workbook()

    # ── Sheet 1: 认缴流量 ──
    ws1 = wb.active
    ws1.title = "认缴流量"
    h1 = ["PDF页码", "增资日期", "认购方", "认购数量(万股)", "认购金额(万元)", "认购价格(元/股)", "原文证据"]
    _header(ws1, h1)
    for r in sub_flows:
        ws1.append([
            r.get("source_page"), r.get("subscription_date"), r.get("subscriber_name"),
            r.get("shares_subscribed"), r.get("amount_subscribed"),
            r.get("price_per_share"), (r.get("evidence_text") or "")[:500],
        ])
    _widths(ws1, [14, 13, 22, 14, 14, 14, 60])

    # ── Sheet 2: 股权存量 ──
    ws2 = wb.create_sheet("股权存量")
    h2 = ["PDF页码", "时点", "股权结构口径", "总股本(万股)", "总出资额(万元)",
          "股东名称", "持股数(万股)", "出资额(万元)", "持股比例", "原文证据"]
    _header(ws2, h2)
    for r in eq_snaps:
        ws2.append([
            r.get("source_page"), r.get("snapshot_date"), r.get("snapshot_type"),
            r.get("total_shares"), r.get("total_capital"),
            r.get("shareholder_name"), r.get("shares_held"),
            r.get("capital_contribution"), r.get("shareholding_ratio"),
            (r.get("evidence_text") or "")[:500],
        ])
    _widths(ws2, [14, 13, 18, 13, 13, 22, 13, 13, 10, 60])

    # ── Sheet 3: schema_cross_check ──
    ws3 = wb.create_sheet("schema_cross_check")

    # 3a: Schema 字段定义
    ws3.merge_cells("A1:D1")
    ws3.cell(row=1, column=1, value="Pydantic Schema 字段定义").font = Font(bold=True, size=13)
    h3a = ["记录类型", "字段名", "类型", "必填"]
    for c, h in enumerate(h3a, 1):
        cell = ws3.cell(row=2, column=c, value=h)
        cell.font = HF; cell.fill = HFILL; cell.alignment = HA; cell.border = BDR
    sf_fields = [
        ("subscription_flow", "source_page", "str", "✓"),
        ("subscription_flow", "subscription_date", "str(YYYY-MM-DD)", "✓"),
        ("subscription_flow", "subscriber_name", "str", "✓"),
        ("subscription_flow", "shares_subscribed", "float|null", ""),
        ("subscription_flow", "amount_subscribed", "float|null", ""),
        ("subscription_flow", "price_per_share", "float|null", ""),
        ("subscription_flow", "evidence_text", "str(min=20)", "✓"),
    ]
    es_fields = [
        ("equity_snapshot", "source_page", "str", "✓"),
        ("equity_snapshot", "snapshot_date", "str(YYYY-MM-DD)", "✓"),
        ("equity_snapshot", "snapshot_type", "str", "✓"),
        ("equity_snapshot", "total_shares", "float|null", ""),
        ("equity_snapshot", "total_capital", "float|null", ""),
        ("equity_snapshot", "shareholder_name", "str", "✓"),
        ("equity_snapshot", "shares_held", "float|null", ""),
        ("equity_snapshot", "capital_contribution", "float|null", ""),
        ("equity_snapshot", "shareholding_ratio", "str|null", ""),
        ("equity_snapshot", "evidence_text", "str(min=20)", "✓"),
    ]
    row = 3
    for fld in sf_fields + es_fields:
        for c, v in enumerate(fld, 1):
            ws3.cell(row=row, column=c, value=v)
        row += 1

    # 3b: Cross-check 结果（完整数字列）
    row += 2
    ws3.merge_cells(f"A{row}:L{row}")
    ws3.cell(row=row, column=1, value="Cross-Check 结果（相邻时点股本追踪 + 认缴存量对应）").font = Font(bold=True, size=13)
    row += 2

    cc_headers = [
        "公司", "上一时点", "当前时点", "股东名称",
        "上一时点持股数\n(万股)", "上一时点出资额\n(万元注册资本)",
        "本次认缴/变化\n(万股)", "预期变更后\n持股数(万股)",
        "PDF披露\n持股数(万股)", "差额(万股)", "状态", "说明"
    ]
    for c, h in enumerate(cc_headers, 1):
        cell = ws3.cell(row=row, column=c, value=h)
        cell.font = HF; cell.fill = HFILL; cell.alignment = HA; cell.border = BDR

    # 读取 cross_check_summary.csv
    cc_path = PROJECT_ROOT / "logs" / "cross_check_summary.csv"
    if cc_path.exists():
        import csv
        with open(cc_path, encoding="utf-8-sig") as fh:
            for r in csv.DictReader(fh):
                row += 1
                ws3.cell(row=row, column=1, value=r.get("company", ""))
                ws3.cell(row=row, column=2, value=r.get("snapshot_date_from", ""))
                ws3.cell(row=row, column=3, value=r.get("snapshot_date_to", ""))
                ws3.cell(row=row, column=4, value=r.get("shareholder_name", ""))
                ws3.cell(row=row, column=5, value=r.get("previous_shares", ""))
                ws3.cell(row=row, column=6, value=r.get("previous_capital", ""))
                ws3.cell(row=row, column=7, value=r.get("subscription_change", ""))
                ws3.cell(row=row, column=8, value=r.get("expected_shares", ""))
                ws3.cell(row=row, column=9, value=r.get("pdf_disclosed_shares", ""))
                ws3.cell(row=row, column=10, value=r.get("difference", ""))
                ws3.cell(row=row, column=11, value=r.get("status", ""))
                ws3.cell(row=row, column=12, value=r.get("detail", ""))

    _widths(ws3, [14, 10, 10, 18, 13, 14, 13, 13, 13, 11, 8, 22])

    # 3c: 汇总统计
    row += 2
    ws3.cell(row=row, column=1, value=f"认缴流量: {len(sub_flows)} 条").font = Font(bold=True)
    row += 1
    ws3.cell(row=row, column=1, value=f"股权存量: {len(eq_snaps)} 条").font = Font(bold=True)
    row += 1
    ws3.cell(row=row, column=1, value=f"生成时间: {datetime.now().isoformat()}")

    wb.save(out_path)


def main():
    if not HAS_XL:
        print("需要 openpyxl: pip install openpyxl"); sys.exit(1)

    print("=" * 60)
    print("JSONL → Excel (三 Sheet: 认缴流量 + 股权存量 + schema_cross_check)")
    print("=" * 60)

    for f in sorted(JSONL_DIR.glob("*.jsonl")):
        out = EXCEL_DIR / f"{f.stem}_三表抽取.xlsx"
        jsonl_to_excel(f, out)
        print(f"  {f.stem} -> {out.name}")

    print(f"\n输出: {EXCEL_DIR}")


if __name__ == "__main__":
    main()
